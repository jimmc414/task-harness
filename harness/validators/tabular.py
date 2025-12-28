"""Tabular file validators for Task Harness.

These validators check CSV and Excel files for headers and row counts.
Uses pandas for reading and openpyxl for Excel support.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from harness.models import ValidationResult
from harness.validators.base import Validator


class TabularFileValid(Validator):
    """Validate that a tabular file (CSV/Excel) has required structure.

    Checks:
    - File exists and is readable
    - Required headers are present (if specified)
    - File has at least min_data_rows of data (excluding header)

    Memory-efficient: Only reads headers and counts rows, doesn't load all data.

    Example:
        preconditions = [
            TabularFileValid("data/input.csv", required_headers=["id", "name", "value"]),
            TabularFileValid("report.xlsx", sheet_name="Data", min_data_rows=10),
        ]
    """

    name = "TabularFileValid"

    def __init__(
        self,
        path: str | Path,
        required_headers: list[str] | None = None,
        min_data_rows: int = 1,
        sheet_name: str | int = 0,
        from_context: bool = False,
    ):
        """Initialize the validator.

        Args:
            path: Path to the file, or context key if from_context is True.
            required_headers: List of column names that must be present.
            min_data_rows: Minimum number of data rows (excluding header).
            sheet_name: For Excel files, sheet name or 0-based index.
            from_context: If True, look up the path from context[path].
        """
        if min_data_rows < 0:
            raise ValueError("min_data_rows cannot be negative")

        self.path = str(path)
        self.required_headers = required_headers or []
        self.min_data_rows = min_data_rows
        self.sheet_name = sheet_name
        self.from_context = from_context

    def check(self, context: dict[str, Any]) -> ValidationResult:
        """Validate the tabular file."""
        try:
            resolved_path = self._resolve_path(self.path, context, self.from_context)
        except (KeyError, ValueError) as e:
            return ValidationResult.failure(self.name, str(e))

        if not resolved_path.exists():
            return ValidationResult.failure(
                self.name,
                f"File not found: {resolved_path}",
            )

        if not resolved_path.is_file():
            return ValidationResult.failure(
                self.name,
                f"Path is not a file: {resolved_path}",
            )

        # Determine file type and read
        suffix = resolved_path.suffix.lower()
        try:
            if suffix in (".xlsx", ".xls", ".xlsm"):
                headers, row_count = self._read_excel(resolved_path)
            elif suffix in (".csv", ".tsv", ".txt"):
                headers, row_count = self._read_csv(resolved_path)
            else:
                return ValidationResult.failure(
                    self.name,
                    f"Unsupported file type: {suffix}",
                )
        except Exception as e:
            return ValidationResult.failure(
                self.name,
                f"Error reading file: {e}",
                details={"path": str(resolved_path), "error": str(e)},
            )

        # Check required headers
        if self.required_headers:
            missing = set(self.required_headers) - set(headers)
            if missing:
                return ValidationResult.failure(
                    self.name,
                    f"Missing required headers: {sorted(missing)}",
                    details={
                        "required": self.required_headers,
                        "found": headers,
                        "missing": sorted(missing),
                    },
                )

        # Check minimum rows
        if row_count < self.min_data_rows:
            return ValidationResult.failure(
                self.name,
                f"Not enough data rows: found {row_count}, need {self.min_data_rows}",
                details={
                    "row_count": row_count,
                    "min_required": self.min_data_rows,
                },
            )

        return ValidationResult.success(
            self.name,
            f"Valid tabular file: {len(headers)} columns, {row_count} rows",
        )

    def _read_csv(self, path: Path) -> tuple[list[str], int]:
        """Read CSV headers and count rows efficiently.

        Tries UTF-8-sig first (handles BOM), falls back to latin-1.
        """
        import pandas as pd

        # Try UTF-8 with BOM handling first
        encodings = ["utf-8-sig", "latin-1"]

        for encoding in encodings:
            try:
                # Read just headers (nrows=0 still reads header row)
                df_headers = pd.read_csv(path, nrows=0, encoding=encoding)
                headers = list(df_headers.columns)

                # Count rows efficiently
                row_count = 0
                with open(path, "r", encoding=encoding, newline="") as f:
                    reader = csv.reader(f)
                    next(reader, None)  # skip header
                    for row in reader:
                        # Count non-empty rows
                        if any(cell.strip() for cell in row):
                            row_count += 1

                return headers, row_count

            except UnicodeDecodeError:
                continue
            except Exception:
                if encoding == encodings[-1]:
                    raise
                continue

        raise ValueError("Could not decode file with any supported encoding")

    def _read_excel(self, path: Path) -> tuple[list[str], int]:
        """Read Excel headers and count rows efficiently."""
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            # Get the worksheet
            if isinstance(self.sheet_name, int):
                if self.sheet_name >= len(wb.worksheets):
                    raise ValueError(
                        f"Sheet index {self.sheet_name} out of range "
                        f"(file has {len(wb.worksheets)} sheets)"
                    )
                sheet = wb.worksheets[self.sheet_name]
            else:
                if self.sheet_name not in wb.sheetnames:
                    raise ValueError(
                        f"Sheet '{self.sheet_name}' not found. "
                        f"Available: {wb.sheetnames}"
                    )
                sheet = wb[self.sheet_name]

            # Get headers from first row
            headers = []
            row_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            first_row = next(row_iter, None)
            if first_row:
                headers = [str(cell) if cell is not None else "" for cell in first_row]

            # Count non-empty data rows
            row_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Count rows with at least one non-empty cell
                if any(cell is not None and str(cell).strip() for cell in row):
                    row_count += 1

            return headers, row_count

        finally:
            wb.close()

    def __repr__(self) -> str:
        parts = [f"{self.path!r}"]
        if self.required_headers:
            parts.append(f"required_headers={self.required_headers!r}")
        if self.min_data_rows != 1:
            parts.append(f"min_data_rows={self.min_data_rows}")
        if self.sheet_name != 0:
            parts.append(f"sheet_name={self.sheet_name!r}")
        if self.from_context:
            parts.append("from_context=True")
        return f"TabularFileValid({', '.join(parts)})"


class TabularFileRowCount(Validator):
    """Check if a tabular file has a row count within a specified range.

    Example:
        preconditions = [
            TabularFileRowCount("data.csv", min_rows=100),
            TabularFileRowCount("data.csv", max_rows=10000),
            TabularFileRowCount("data.xlsx", min_rows=10, max_rows=1000),
        ]
    """

    name = "TabularFileRowCount"

    def __init__(
        self,
        path: str | Path,
        min_rows: int = 0,
        max_rows: int | None = None,
        sheet_name: str | int = 0,
        from_context: bool = False,
    ):
        """Initialize the validator.

        Args:
            path: Path to the file, or context key if from_context is True.
            min_rows: Minimum number of data rows (default 0).
            max_rows: Maximum number of data rows (None = no limit).
            sheet_name: For Excel files, sheet name or 0-based index.
            from_context: If True, look up the path from context[path].
        """
        if min_rows < 0:
            raise ValueError("min_rows cannot be negative")
        if max_rows is not None and max_rows < min_rows:
            raise ValueError("max_rows cannot be less than min_rows")

        self.path = str(path)
        self.min_rows = min_rows
        self.max_rows = max_rows
        self.sheet_name = sheet_name
        self.from_context = from_context

    def check(self, context: dict[str, Any]) -> ValidationResult:
        """Check if the file has the expected row count."""
        # Use TabularFileValid to get row count
        inner = TabularFileValid(
            self.path,
            min_data_rows=0,  # We'll check ourselves
            sheet_name=self.sheet_name,
            from_context=self.from_context,
        )

        try:
            resolved_path = inner._resolve_path(self.path, context, self.from_context)
        except (KeyError, ValueError) as e:
            return ValidationResult.failure(self.name, str(e))

        if not resolved_path.exists():
            return ValidationResult.failure(
                self.name,
                f"File not found: {resolved_path}",
            )

        suffix = resolved_path.suffix.lower()
        try:
            if suffix in (".xlsx", ".xls", ".xlsm"):
                _, row_count = inner._read_excel(resolved_path)
            elif suffix in (".csv", ".tsv", ".txt"):
                _, row_count = inner._read_csv(resolved_path)
            else:
                return ValidationResult.failure(
                    self.name,
                    f"Unsupported file type: {suffix}",
                )
        except Exception as e:
            return ValidationResult.failure(
                self.name,
                f"Error reading file: {e}",
            )

        if row_count < self.min_rows:
            return ValidationResult.failure(
                self.name,
                f"Too few rows: {row_count} < {self.min_rows}",
                details={"row_count": row_count, "min_rows": self.min_rows},
            )

        if self.max_rows is not None and row_count > self.max_rows:
            return ValidationResult.failure(
                self.name,
                f"Too many rows: {row_count} > {self.max_rows}",
                details={"row_count": row_count, "max_rows": self.max_rows},
            )

        return ValidationResult.success(
            self.name,
            f"Row count {row_count} is within range",
        )

    def __repr__(self) -> str:
        parts = [f"{self.path!r}"]
        if self.min_rows > 0:
            parts.append(f"min_rows={self.min_rows}")
        if self.max_rows is not None:
            parts.append(f"max_rows={self.max_rows}")
        if self.sheet_name != 0:
            parts.append(f"sheet_name={self.sheet_name!r}")
        if self.from_context:
            parts.append("from_context=True")
        return f"TabularFileRowCount({', '.join(parts)})"
